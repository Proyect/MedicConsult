"""
Módulo para generar reportes y exportaciones
"""
import io
from datetime import datetime, timedelta
from django.http import HttpResponse
from django.db.models import Count, Q
from django.template.loader import render_to_string
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from .models import Person, Doctor, Consult, MedicalRecord, UserProfile

class ReportGenerator:
    """Generador de reportes en PDF y Excel"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.setup_custom_styles()
    
    def setup_custom_styles(self):
        """Configurar estilos personalizados para PDF"""
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,  # Centrado
            textColor=colors.darkblue
        )
        
        self.subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=12,
            spaceAfter=12,
            textColor=colors.darkblue
        )
        
        self.normal_style = ParagraphStyle(
            'CustomNormal',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=6
        )

    def generate_patients_pdf(self, patients, title="Reporte de Pacientes"):
        """Generar reporte de pacientes en PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Contenido del documento
        story = []
        
        # Título
        story.append(Paragraph(title, self.title_style))
        story.append(Spacer(1, 12))
        
        # Información del reporte
        story.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}", self.normal_style))
        story.append(Paragraph(f"Total de pacientes: {patients.count()}", self.normal_style))
        story.append(Spacer(1, 20))
        
        # Tabla de pacientes
        if patients.exists():
            data = [['Nombre', 'Apellido', 'DNI', 'Edad', 'Teléfono', 'Email']]
            
            for patient in patients:
                data.append([
                    patient.name,
                    patient.last_name,
                    patient.dni,
                    str(patient.age),
                    patient.phone,
                    patient.email
                ])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        else:
            story.append(Paragraph("No se encontraron pacientes con los criterios especificados.", self.normal_style))
        
        doc.build(story)
        buffer.seek(0)
        return buffer

    def generate_consults_pdf(self, consults, title="Reporte de Consultas"):
        """Generar reporte de consultas en PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        story = []
        
        # Título
        story.append(Paragraph(title, self.title_style))
        story.append(Spacer(1, 12))
        
        # Información del reporte
        story.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}", self.normal_style))
        story.append(Paragraph(f"Total de consultas: {consults.count()}", self.normal_style))
        story.append(Spacer(1, 20))
        
        # Tabla de consultas
        if consults.exists():
            data = [['Fecha', 'Paciente', 'Doctor', 'Tipo', 'Motivo']]
            
            for consult in consults:
                data.append([
                    consult.date.strftime('%d/%m/%Y %H:%M'),
                    f"{consult.patient.name} {consult.patient.last_name}",
                    f"Dr. {consult.doctor.full_name}",
                    consult.get_consult_type_display(),
                    consult.reason[:50] + "..." if len(consult.reason) > 50 else consult.reason
                ])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        else:
            story.append(Paragraph("No se encontraron consultas con los criterios especificados.", self.normal_style))
        
        doc.build(story)
        buffer.seek(0)
        return buffer

    def generate_statistics_pdf(self, title="Estadísticas Generales"):
        """Generar reporte de estadísticas en PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        story = []
        
        # Título
        story.append(Paragraph(title, self.title_style))
        story.append(Spacer(1, 12))
        
        # Estadísticas generales
        total_patients = Person.objects.filter(is_active=True).count()
        total_doctors = Doctor.objects.filter(is_active=True).count()
        total_consults = Consult.objects.count()
        
        # Consultas por mes (últimos 6 meses)
        six_months_ago = datetime.now() - timedelta(days=180)
        consults_by_month = Consult.objects.filter(date__gte=six_months_ago).extra(
            select={'month': "EXTRACT(month FROM date)"}
        ).values('month').annotate(count=Count('id')).order_by('month')
        
        # Consultas por tipo
        consults_by_type = Consult.objects.values('consult_type').annotate(count=Count('id'))
        
        story.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}", self.normal_style))
        story.append(Spacer(1, 20))
        
        # Resumen general
        story.append(Paragraph("Resumen General", self.subtitle_style))
        summary_data = [
            ['Métrica', 'Valor'],
            ['Total de Pacientes', str(total_patients)],
            ['Total de Doctores', str(total_doctors)],
            ['Total de Consultas', str(total_consults)],
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Consultas por tipo
        if consults_by_type:
            story.append(Paragraph("Consultas por Tipo", self.subtitle_style))
            type_data = [['Tipo de Consulta', 'Cantidad']]
            
            for item in consults_by_type:
                consult_type_display = dict(Consult.CONSULT_TYPE_CHOICES).get(item['consult_type'], item['consult_type'])
                type_data.append([consult_type_display, str(item['count'])])
            
            type_table = Table(type_data)
            type_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(type_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer

    def generate_patients_excel(self, patients, title="Reporte de Pacientes"):
        """Generar reporte de pacientes en Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pacientes"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Título
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = center_alignment
        
        # Información del reporte
        ws['A3'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A4'] = f"Total de pacientes: {patients.count()}"
        
        # Encabezados
        headers = ['Nombre', 'Apellido', 'DNI', 'Edad', 'Teléfono', 'Email', 'Dirección']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Datos
        row = 7
        for patient in patients:
            ws.cell(row=row, column=1, value=patient.name)
            ws.cell(row=row, column=2, value=patient.last_name)
            ws.cell(row=row, column=3, value=patient.dni)
            ws.cell(row=row, column=4, value=patient.age)
            ws.cell(row=row, column=5, value=patient.phone)
            ws.cell(row=row, column=6, value=patient.email)
            ws.cell(row=row, column=7, value=patient.address)
            row += 1
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return wb

    def generate_consults_excel(self, consults, title="Reporte de Consultas"):
        """Generar reporte de consultas en Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consultas"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Título
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = center_alignment
        
        # Información del reporte
        ws['A3'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A4'] = f"Total de consultas: {consults.count()}"
        
        # Encabezados
        headers = ['Fecha', 'Paciente', 'Doctor', 'Tipo', 'Motivo']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Datos
        row = 7
        for consult in consults:
            ws.cell(row=row, column=1, value=consult.date.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=2, value=f"{consult.patient.name} {consult.patient.last_name}")
            ws.cell(row=row, column=3, value=f"Dr. {consult.doctor.full_name}")
            ws.cell(row=row, column=4, value=consult.get_consult_type_display())
            ws.cell(row=row, column=5, value=consult.reason)
            row += 1
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return wb

def get_statistics_data():
    """Obtener datos estadísticos para el dashboard"""
    from django.db.models import Count, Q
    from datetime import datetime, timedelta
    
    # Estadísticas básicas
    total_patients = Person.objects.filter(is_active=True).count()
    total_doctors = Doctor.objects.filter(is_active=True).count()
    total_consults = Consult.objects.count()
    
    # Consultas del mes actual
    current_month = datetime.now().replace(day=1)
    consults_this_month = Consult.objects.filter(date__gte=current_month).count()
    
    # Consultas por tipo
    consults_by_type = list(Consult.objects.values('consult_type').annotate(count=Count('id')))
    
    # Consultas por doctor (últimos 30 días)
    thirty_days_ago = datetime.now() - timedelta(days=30)
    consults_by_doctor = list(Consult.objects.filter(date__gte=thirty_days_ago)
                             .values('doctor__user__first_name', 'doctor__user__last_name')
                             .annotate(count=Count('id'))
                             .order_by('-count')[:5])
    
    # Pacientes por género
    patients_by_gender = list(Person.objects.filter(is_active=True)
                             .values('gender')
                             .annotate(count=Count('id')))
    
    return {
        'total_patients': total_patients,
        'total_doctors': total_doctors,
        'total_consults': total_consults,
        'consults_this_month': consults_this_month,
        'consults_by_type': consults_by_type,
        'consults_by_doctor': consults_by_doctor,
        'patients_by_gender': patients_by_gender,
    }

